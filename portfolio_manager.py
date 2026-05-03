# -*- coding: utf-8 -*-
"""
基金组合管理模块
提供组合 CRUD、风险评估、基金推荐、历史回测、名称生成等功能。
依赖：app.py 的数据库/缓存工具函数，fund_analyzer.py 的 FundScreener/ReportGenerator，
     fund_crawler.py 的数据获取函数。
"""
import json
import random
import concurrent.futures
import threading
import os
from datetime import datetime, date, timedelta
from io import BytesIO

import pandas as pd
import numpy as np
from flask import request, jsonify, session

# ── 延迟导入以避免循环依赖 ────────────────────────────────────
_get_mysql_pool = None
_get_cache = None
_set_cache = None
_generate_cache_key = None
_REDIS_AVAILABLE = False
_SQLITE_DB_PATH = None

def _init_deps():
    """初始化对 app.py 的依赖引用"""
    global _get_mysql_pool, _get_cache, _set_cache, _generate_cache_key
    global _REDIS_AVAILABLE, _SQLITE_DB_PATH
    import app
    _get_mysql_pool = app.get_mysql_pool
    _get_cache = app.get_cache
    _set_cache = app.set_cache
    _generate_cache_key = app.generate_cache_key
    _REDIS_AVAILABLE = app.REDIS_AVAILABLE
    _SQLITE_DB_PATH = app.SQLITE_DB_PATH


# ══════════════════════════════════════════════════════════════
# 常量
# ══════════════════════════════════════════════════════════════

RISK_LEVEL_MAP = {
    # 每题20-80分，5题满分400。映射回PRD的百分制区间。
    (100, 160): ('保守型', 5.0),
    (161, 240): ('稳健型', 10.0),
    (241, 300): ('平衡型', 15.0),
    (301, 340): ('成长型', 25.0),
    (341, 400): ('激进型', 35.0),
}

RISK_QUESTIONS = [
    {
        'id': 1,
        'question': '您的投资经验有多久？',
        'options': [
            {'text': '0-1年', 'score': 20},
            {'text': '1-3年', 'score': 40},
            {'text': '3-5年', 'score': 60},
            {'text': '5年以上', 'score': 80},
        ]
    },
    {
        'id': 2,
        'question': '您的可投资资金占家庭总资产的比例是多少？',
        'options': [
            {'text': '<10%', 'score': 20},
            {'text': '10%-30%', 'score': 40},
            {'text': '30%-50%', 'score': 60},
            {'text': '≥50%', 'score': 80},
        ]
    },
    {
        'id': 3,
        'question': '您计划的投资期限是多久？',
        'options': [
            {'text': '<1年', 'score': 20},
            {'text': '1-3年', 'score': 40},
            {'text': '3-5年', 'score': 60},
            {'text': '5年以上', 'score': 80},
        ]
    },
    {
        'id': 4,
        'question': '如果您的投资在短期内亏损了20%，您会怎么做？',
        'options': [
            {'text': '立即卖出，避免更大损失', 'score': 20},
            {'text': '卖出部分，降低风险', 'score': 40},
            {'text': '继续持有，等待反弹', 'score': 60},
            {'text': '加仓买入，摊低成本', 'score': 80},
        ]
    },
    {
        'id': 5,
        'question': '您的投资目标是什么？',
        'options': [
            {'text': '保本增值，尽可能避免亏损', 'score': 20},
            {'text': '稳健收益，接受小幅波动', 'score': 40},
            {'text': '追求较高收益，接受中等波动', 'score': 60},
            {'text': '追求高收益，接受较大波动', 'score': 80},
        ]
    }
]

# 资产配置比例（按风险等级）
ASSET_ALLOCATION = {
    '保守型': {'债券型': 55, '货币型': 25, '混合型': 15, '股票型': 5},
    '稳健型': {'债券型': 40, '混合型': 30, '股票型': 15, '货币型': 10, '指数型': 5},
    '平衡型': {'混合型': 35, '股票型': 25, '债券型': 20, '指数型': 15, '货币型': 5},
    '成长型': {'股票型': 40, '混合型': 25, '指数型': 20, '债券型': 10, '货币型': 5},
    '激进型': {'股票型': 50, '指数型': 25, '混合型': 15, '债券型': 5, '货币型': 5},
}

# 文艺名称库
PORTFOLIO_NAMES = {
    '保守型': {
        '自然意象': ['静水流深', '稳若磐石', '和风细雨', '松柏长青', '青山不老', '幽谷清泉',
                     '晨曦微露', '秋水长天', '寒梅傲雪', '竹影清风'],
        '古典诗词': ['行稳致远', '厚德载物', '上善若水', '大巧若拙', '抱朴守拙',
                     '宁静致远', '温故知新', '履霜坚冰', '卑以自牧', '安之若素'],
        '投资理念': ['本金为先', '稳字当头', '安睡投资', '保守致远', '安全边际',
                     '价值坚守', '底线思维', '风险可控', '持盈保泰', '稳健前行'],
    },
    '稳健型': {
        '自然意象': ['春华秋实', '月华如水', '碧波万顷', '云淡风轻', '瑞雪兆丰',
                     '映日荷花', '梧桐栖凤', '山高水长', '海纳百川', '林深见鹿'],
        '古典诗词': ['积健为雄', '博观约取', '细水长流', '厚积薄发', '循序渐进',
                     '源远流长', '百川归海', '玉汝于成', '跬步千里', '日升月恒'],
        '投资理念': ['稳健增长', '均衡配置', '攻守兼备', '复利魔力', '长期主义',
                     '资产保值', '稳步向前', '平衡致胜', '守正出奇', '价值成长'],
    },
    '平衡型': {
        '自然意象': ['鲲鹏展翅', '龙跃云津', '骏马奔腾', '鹰击长空', '虎啸风生',
                     '凤鸣岐山', '鱼跃龙门', '鹤立鸡群', '鹏程万里', '龙腾四海'],
        '古典诗词': ['乘风破浪', '中流击楫', '攀云追月', '登高望远', '凌云之志',
                     '志在千里', '气贯长虹', '格物致知', '融会贯通', '高瞻远瞩'],
        '投资理念': ['进可攻退可守', '资产再平衡', '动态调整', '灵活配置', '顺势而为',
                     '攻守平衡', '稳中求进', '多元分散', '战略配置', '战术灵活'],
    },
    '成长型': {
        '自然意象': ['旭日东升', '星火燎原', '破茧成蝶', '鹰隼试翼', '大鹏展翅',
                     '乘风破浪', '奔腾入海', '扶摇直上', '驰骋天地', '一飞冲天'],
        '古典诗词': ['鸿鹄高飞', '不鸣则已', '一日千里', '青云直上', '长风破浪',
                     '会当凌绝', '披荆斩棘', '勇往直前', '开疆拓土', '锐意进取'],
        '投资理念': ['成长驱动', '长期复利', '价值发现', '趋势为王', '优选标的',
                     '核心增长', '掘金时代', '新兴赛道', '增长红利', '创新引擎'],
    },
    '激进型': {
        '自然意象': ['烈焰燎原', '雷霆万钧', '狂风骤雨', '惊涛骇浪', '火山喷薄',
                     '星陨如雨', '电光火石', '怒海争锋', '风暴之眼', '赤焰焚天'],
        '古典诗词': ['破釜沉舟', '背水一战', '剑走偏锋', '出奇制胜', '先发制人',
                     '一骑绝尘', '势如破竹', '独辟蹊径', '快马加鞭', '雷霆出击'],
        '投资理念': ['激进增值', '高回报策略', '趋势加速', '杠杆效应', '超额收益',
                     '主动出击', '集中火力', '效率优先', '弯道超车', '突破瓶颈'],
    },
}


# ══════════════════════════════════════════════════════════════
# 数据库初始化
# ══════════════════════════════════════════════════════════════

_MYSQL_TABLES = {
    'portfolio_settings': '''
        CREATE TABLE IF NOT EXISTS portfolio_settings (
            id INT PRIMARY KEY DEFAULT 1,
            payment_enabled TINYINT(1) DEFAULT 0,
            free_trial_days INT DEFAULT 7,
            monthly_price DECIMAL(10,2) DEFAULT 19.90,
            quarterly_price DECIMAL(10,2) DEFAULT 49.90,
            annual_price DECIMAL(10,2) DEFAULT 169.00,
            user_payment_enabled TINYINT(1) DEFAULT 1,
            user_free_access TINYINT(1) DEFAULT 1,
            trial_start_date DATETIME,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    ''',
    'payment_records': '''
        CREATE TABLE IF NOT EXISTS payment_records (
            id INT AUTO_INCREMENT PRIMARY KEY,
            plan_type VARCHAR(20) NOT NULL,
            amount DECIMAL(10,2) NOT NULL,
            payment_method VARCHAR(20) DEFAULT '模拟支付',
            status VARCHAR(20) DEFAULT 'paid',
            start_date DATETIME NOT NULL,
            end_date DATETIME NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    ''',
    'risk_assessments': '''
        CREATE TABLE IF NOT EXISTS risk_assessments (
            id INT AUTO_INCREMENT PRIMARY KEY,
            user_id INT NOT NULL,
            total_score INT NOT NULL,
            risk_level VARCHAR(20) NOT NULL,
            max_drawdown DECIMAL(5,2) NOT NULL,
            manual_override TINYINT(1) DEFAULT 0,
            manual_max_drawdown DECIMAL(5,2) DEFAULT NULL,
            risk_warning_accepted TINYINT(1) DEFAULT 0,
            assessment_date DATETIME NOT NULL,
            expiry_date DATETIME NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    ''',
    'portfolios': '''
        CREATE TABLE IF NOT EXISTS portfolios (
            id INT AUTO_INCREMENT PRIMARY KEY,
            user_id INT NOT NULL,
            name VARCHAR(100) NOT NULL,
            description TEXT,
            tags VARCHAR(255),
            risk_level VARCHAR(20) NOT NULL DEFAULT '平衡型',
            target_max_drawdown DECIMAL(5,2),
            created_from VARCHAR(20) DEFAULT 'custom',
            is_active TINYINT(1) DEFAULT 1,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    ''',
    'portfolio_holdings': '''
        CREATE TABLE IF NOT EXISTS portfolio_holdings (
            id INT AUTO_INCREMENT PRIMARY KEY,
            portfolio_id INT NOT NULL,
            fund_code VARCHAR(10) NOT NULL,
            fund_name VARCHAR(100),
            fund_type VARCHAR(30),
            weight DECIMAL(5,2) NOT NULL,
            added_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE KEY uk_portfolio_fund (portfolio_id, fund_code)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    ''',
    'portfolio_nav_cache': '''
        CREATE TABLE IF NOT EXISTS portfolio_nav_cache (
            id INT AUTO_INCREMENT PRIMARY KEY,
            portfolio_id INT NOT NULL,
            nav_date DATE NOT NULL,
            nav_value DECIMAL(10,4) NOT NULL,
            UNIQUE KEY uk_portfolio_date (portfolio_id, nav_date),
            KEY idx_portfolio_id (portfolio_id)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    ''',
    'portfolio_backtest_cache': '''
        CREATE TABLE IF NOT EXISTS portfolio_backtest_cache (
            id INT AUTO_INCREMENT PRIMARY KEY,
            portfolio_id INT NOT NULL,
            start_date DATE NOT NULL,
            end_date DATE NOT NULL,
            backtest_data LONGTEXT NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE KEY uk_backtest (portfolio_id, start_date, end_date)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    ''',
}

_SQLITE_TABLES = {
    'portfolio_settings': '''
        CREATE TABLE IF NOT EXISTS portfolio_settings (
            id INTEGER PRIMARY KEY DEFAULT 1,
            payment_enabled INTEGER DEFAULT 0,
            free_trial_days INTEGER DEFAULT 7,
            monthly_price REAL DEFAULT 19.90,
            quarterly_price REAL DEFAULT 49.90,
            annual_price REAL DEFAULT 169.00,
            user_payment_enabled INTEGER DEFAULT 1,
            user_free_access INTEGER DEFAULT 1,
            trial_start_date TEXT,
            updated_at TEXT
        )
    ''',
    'payment_records': '''
        CREATE TABLE IF NOT EXISTS payment_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            plan_type TEXT NOT NULL,
            amount REAL NOT NULL,
            payment_method TEXT DEFAULT '模拟支付',
            status TEXT DEFAULT 'paid',
            start_date TEXT NOT NULL,
            end_date TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )
    ''',
    'risk_assessments': '''
        CREATE TABLE IF NOT EXISTS risk_assessments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            total_score INTEGER NOT NULL,
            risk_level TEXT NOT NULL,
            max_drawdown REAL NOT NULL,
            manual_override INTEGER DEFAULT 0,
            manual_max_drawdown REAL,
            risk_warning_accepted INTEGER DEFAULT 0,
            assessment_date TEXT NOT NULL,
            expiry_date TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )
    ''',
    'portfolios': '''
        CREATE TABLE IF NOT EXISTS portfolios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            description TEXT,
            tags TEXT,
            risk_level TEXT NOT NULL DEFAULT '平衡型',
            target_max_drawdown REAL,
            created_from TEXT DEFAULT 'custom',
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        )
    ''',
    'portfolio_holdings': '''
        CREATE TABLE IF NOT EXISTS portfolio_holdings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            portfolio_id INTEGER NOT NULL,
            fund_code TEXT NOT NULL,
            fund_name TEXT,
            fund_type TEXT,
            weight REAL NOT NULL,
            added_at TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(portfolio_id, fund_code)
        )
    ''',
    'portfolio_nav_cache': '''
        CREATE TABLE IF NOT EXISTS portfolio_nav_cache (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            portfolio_id INTEGER NOT NULL,
            nav_date TEXT NOT NULL,
            nav_value REAL NOT NULL,
            UNIQUE(portfolio_id, nav_date)
        )
    ''',
    'portfolio_backtest_cache': '''
        CREATE TABLE IF NOT EXISTS portfolio_backtest_cache (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            portfolio_id INTEGER NOT NULL,
            start_date TEXT NOT NULL,
            end_date TEXT NOT NULL,
            backtest_data TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(portfolio_id, start_date, end_date)
        )
    ''',
}


def _ensure_deps():
    if _get_mysql_pool is None:
        _init_deps()


def _db_execute(sql, params=None, fetch=True):
    """执行 MySQL 查询，自动降级到 SQLite"""
    _ensure_deps()
    pool = _get_mysql_pool()
    if pool is not None:
        conn = None
        try:
            conn = pool.get_connection()
            cur = conn.cursor()
            cur.execute(sql, params or ())
            if fetch and sql.strip().upper().startswith('SELECT'):
                rows = cur.fetchall()
            elif fetch:
                conn.commit()
                rows = cur.lastrowid
            else:
                conn.commit()
                rows = cur.lastrowid
            cur.close()
            return rows
        except Exception as e:
            print(f"[portfolio] MySQL error: {e}, falling back to SQLite")
        finally:
            if conn:
                try:
                    conn.close()
                except Exception:
                    pass
    return _sqlite_execute(sql, params, fetch)


def _sqlite_execute(sql, params=None, fetch=True):
    import sqlite3
    sql = sql.replace('%s', '?')
    sql = sql.replace('ON UPDATE CURRENT_TIMESTAMP', '')
    sql = sql.replace('AUTO_INCREMENT', 'AUTOINCREMENT')
    sql = sql.replace('LONGTEXT', 'TEXT')
    sql = sql.replace('TINYINT(1)', 'INTEGER')
    sql = sql.replace('DECIMAL(10,2)', 'REAL')
    sql = sql.replace('DECIMAL(5,2)', 'REAL')
    sql = sql.replace('DECIMAL(10,4)', 'REAL')
    sql = sql.replace('ENGINE=InnoDB DEFAULT CHARSET=utf8mb4', '')
    conn = sqlite3.connect(_SQLITE_DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        cur = conn.cursor()
        cur.execute(sql, params or ())
        if fetch and sql.strip().upper().startswith('SELECT'):
            rows = [dict(r) for r in cur.fetchall()]
        elif fetch and ('INSERT' in sql.upper() or 'UPDATE' in sql.upper() or 'DELETE' in sql.upper()):
            conn.commit()
            rows = cur.lastrowid
        else:
            conn.commit()
            rows = None
        cur.close()
        return rows
    finally:
        conn.close()


def init_portfolio_tables():
    """初始化所有组合模块需要的数据库表（MySQL + SQLite双写）"""
    _ensure_deps()
    import sqlite3 as _sq
    # 先在 MySQL 创建
    for table_name, mysql_sql in _MYSQL_TABLES.items():
        try:
            _db_execute(mysql_sql, fetch=False)
            print(f"[portfolio] MySQL table '{table_name}' OK")
        except Exception as e:
            print(f"[portfolio] MySQL table '{table_name}' failed: {e}")
    # 迁移：为已有表添加 user_id 列（如果不存在）
    _migrate_add_user_id_column()
    # 再在 SQLite 创建（确保 fallback 可用）
    if _SQLITE_DB_PATH:
        try:
            conn = _sq.connect(_SQLITE_DB_PATH)
            for table_name, sqlite_sql in _SQLITE_TABLES.items():
                try:
                    conn.execute(sqlite_sql)
                    conn.commit()
                except Exception as e:
                    print(f"[portfolio] SQLite table '{table_name}' failed: {e}")
            conn.close()
            print("[portfolio] SQLite tables initialized")
        except Exception as e:
            print(f"[portfolio] SQLite init failed: {e}")
    _init_default_settings()


def _migrate_add_user_id_column():
    """为已存在的表添加 user_id 列（兼容已有数据）"""
    import sqlite3 as _sq
    # MySQL 迁移
    try:
        pool = _get_mysql_pool()
        if pool:
            conn = pool.get_connection()
            cur = conn.cursor()
            # portfolios 表
            try:
                cur.execute("ALTER TABLE portfolios ADD COLUMN user_id INT NOT NULL DEFAULT 1")
                conn.commit()
                print("[portfolio] Migrated portfolios.user_id column")
            except Exception:
                pass  # column already exists
            conn.close()
    except Exception:
        pass
    # SQLite 迁移
    if _SQLITE_DB_PATH:
        try:
            conn = _sq.connect(_SQLITE_DB_PATH)
            try:
                conn.execute("ALTER TABLE portfolios ADD COLUMN user_id INTEGER NOT NULL DEFAULT 1")
                conn.commit()
                print("[portfolio] SQLite migrated portfolios.user_id column")
            except Exception:
                pass
            conn.close()
        except Exception:
            pass


def _init_default_settings():
    """确保 portfolio_settings 有默认行"""
    existing = _db_execute("SELECT id FROM portfolio_settings WHERE id=1", fetch=True)
    if not existing or len(existing) == 0:
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        _db_execute(
            "INSERT INTO portfolio_settings (id, payment_enabled, free_trial_days, "
            "monthly_price, quarterly_price, annual_price, user_payment_enabled, "
            "user_free_access, trial_start_date, updated_at) "
            "VALUES (1, 1, 7, 19.90, 49.90, 169.00, 0, 0, %s, %s)",
            (now, now), fetch=False)


# ══════════════════════════════════════════════════════════════
# 设置与支付
# ══════════════════════════════════════════════════════════════

def get_settings():
    rows = _db_execute("SELECT * FROM portfolio_settings WHERE id=1", fetch=True)
    if rows and len(rows) > 0:
        row = rows[0]
        return {
            'payment_enabled': bool(row.get('payment_enabled', 0)),
            'free_trial_days': row.get('free_trial_days', 7),
            'monthly_price': float(row.get('monthly_price', 19.90)),
            'quarterly_price': float(row.get('quarterly_price', 49.90)),
            'annual_price': float(row.get('annual_price', 169.00)),
            'user_payment_enabled': bool(row.get('user_payment_enabled', 1)),
            'user_free_access': bool(row.get('user_free_access', 1)),
            'trial_start_date': str(row.get('trial_start_date', '')) if row.get('trial_start_date') else None,
        }
    return {
        'payment_enabled': True, 'free_trial_days': 7,
        'monthly_price': 19.90, 'quarterly_price': 49.90, 'annual_price': 169.00,
        'user_payment_enabled': False, 'user_free_access': False,
        'trial_start_date': None,
    }


def update_settings(data):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    fields = []
    values = []
    allowed = ['payment_enabled', 'free_trial_days', 'monthly_price', 'quarterly_price',
               'annual_price', 'user_payment_enabled', 'user_free_access', 'trial_start_date']
    for key in allowed:
        if key in data:
            fields.append(f"{key}=%s")
            val = data[key]
            if isinstance(val, bool):
                values.append(1 if val else 0)
            elif val is None:
                values.append(None)
            else:
                values.append(val)
    if fields:
        fields.append("updated_at=%s")
        values.append(now)
        values.append(1)
        _db_execute(f"UPDATE portfolio_settings SET {', '.join(fields)} WHERE id=%s", values, fetch=False)
    return get_settings()


def check_access():
    """检查用户是否可以访问组合模块。返回 (allowed: bool, reason: str)"""
    settings = get_settings()
    if not settings['payment_enabled']:
        return (True, '免费模式已启用')
    if settings['user_free_access']:
        return (True, '用户在白名单中')
    if settings['user_payment_enabled']:
        return (True, '用户已付费')
    trial_start = settings.get('trial_start_date')
    if trial_start:
        try:
            ts = datetime.strptime(str(trial_start)[:10], '%Y-%m-%d')
            days_used = (datetime.now() - ts).days
            if days_used <= settings['free_trial_days']:
                return (True, f'免费试用中（剩余{settings["free_trial_days"] - days_used}天）')
        except Exception:
            pass
    return (False, '需要订阅才能使用此功能')


def get_payment_history():
    rows = _db_execute(
        "SELECT * FROM payment_records ORDER BY created_at DESC LIMIT 20", fetch=True)
    if not rows:
        return []
    result = []
    for r in rows:
        result.append({
            'id': r['id'],
            'plan_type': r['plan_type'],
            'amount': float(r['amount']),
            'payment_method': r.get('payment_method', '模拟支付'),
            'status': r.get('status', 'paid'),
            'start_date': str(r['start_date']),
            'end_date': str(r['end_date']),
            'created_at': str(r.get('created_at', '')),
        })
    return result


def create_subscription(plan_type):
    plans = {'monthly': (19.90, 30), 'quarterly': (49.90, 90), 'annual': (169.00, 365)}
    if plan_type not in plans:
        return {'success': False, 'error': f'无效的套餐类型: {plan_type}'}
    amount, days = plans[plan_type]
    now = datetime.now()
    end_date = now + timedelta(days=days)
    start_str = now.strftime('%Y-%m-%d %H:%M:%S')
    end_str = end_date.strftime('%Y-%m-%d %H:%M:%S')
    _db_execute(
        "INSERT INTO payment_records (plan_type, amount, payment_method, status, start_date, end_date) "
        "VALUES (%s, %s, '模拟支付', 'paid', %s, %s)",
        (plan_type, amount, start_str, end_str), fetch=False)
    _db_execute(
        "UPDATE portfolio_settings SET user_payment_enabled=1, updated_at=%s WHERE id=1",
        (start_str,), fetch=False)
    return {
        'success': True,
        'plan_type': plan_type,
        'amount': amount,
        'end_date': end_str,
        'message': f'订阅成功！{plan_type}套餐，有效期至 {end_str[:10]}',
    }


# ══════════════════════════════════════════════════════════════
# 风险评估
# ══════════════════════════════════════════════════════════════

def get_risk_status(user_id):
    """获取当前有效的风险评估状态"""
    rows = _db_execute(
        "SELECT * FROM risk_assessments WHERE user_id=%s AND expiry_date > %s "
        "ORDER BY created_at DESC LIMIT 1",
        (user_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S')), fetch=True)
    if rows and len(rows) > 0:
        r = rows[0]
        return {
            'has_valid': True,
            'risk_level': r['risk_level'],
            'max_drawdown': float(r['manual_max_drawdown']) if r.get('manual_max_drawdown') else float(r['max_drawdown']),
            'total_score': r['total_score'],
            'manual_override': bool(r.get('manual_override', 0)),
            'assessment_date': str(r['assessment_date']),
            'expiry_date': str(r['expiry_date']),
        }
    return {'has_valid': False, 'risk_level': None, 'max_drawdown': None}


def submit_risk_assessment(answers, user_id):
    """提交风险评估答案，返回风险等级"""
    if not answers or len(answers) != 5:
        return {'success': False, 'error': '需要回答全部5道题目'}
    total_score = sum(int(a) for a in answers)
    risk_level, max_drawdown = _score_to_risk(total_score)
    now = datetime.now()
    expiry = now + timedelta(days=365)
    _db_execute(
        "INSERT INTO risk_assessments (user_id, total_score, risk_level, max_drawdown, "
        "assessment_date, expiry_date) VALUES (%s, %s, %s, %s, %s, %s)",
        (user_id, total_score, risk_level, max_drawdown,
         now.strftime('%Y-%m-%d %H:%M:%S'),
         expiry.strftime('%Y-%m-%d %H:%M:%S')), fetch=False)
    return {
        'success': True,
        'total_score': total_score,
        'risk_level': risk_level,
        'max_drawdown': max_drawdown,
        'expiry_date': expiry.strftime('%Y-%m-%d'),
    }


def override_max_drawdown(new_drawdown, accepted_risk=False, user_id=None):
    status = get_risk_status(user_id)
    if not status['has_valid']:
        return {'success': False, 'error': '请先完成风险评估'}
    if new_drawdown > 35:
        return {'success': False, 'error': '最大回撤不能超过35%'}
    if new_drawdown < 1:
        return {'success': False, 'error': '最大回撤不能低于1%'}
    if not accepted_risk:
        return {'success': False, 'error': '请确认风险提示', 'require_confirmation': True}
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    _db_execute(
        "UPDATE risk_assessments SET manual_override=1, manual_max_drawdown=%s, "
        "risk_warning_accepted=1 WHERE user_id=%s AND expiry_date > %s",
        (new_drawdown, user_id, now), fetch=False)
    return {'success': True, 'max_drawdown': new_drawdown, 'message': '最大回撤已更新'}


def _score_to_risk(score):
    for (lo, hi), (level, dd) in RISK_LEVEL_MAP.items():
        if lo <= score <= hi:
            return level, dd
    return '平衡型', 15.0


def get_risk_history(user_id):
    rows = _db_execute(
        "SELECT * FROM risk_assessments WHERE user_id=%s ORDER BY created_at DESC LIMIT 10",
        (user_id,), fetch=True)
    if not rows:
        return []
    return [{
        'id': r['id'], 'total_score': r['total_score'],
        'risk_level': r['risk_level'],
        'max_drawdown': float(r['manual_max_drawdown']) if r.get('manual_max_drawdown') else float(r['max_drawdown']),
        'manual_override': bool(r.get('manual_override', 0)),
        'assessment_date': str(r['assessment_date']),
        'expiry_date': str(r['expiry_date']),
    } for r in rows]


# ══════════════════════════════════════════════════════════════
# 名称生成器
# ══════════════════════════════════════════════════════════════

def generate_portfolio_names(risk_level='平衡型', style=None):
    names_dict = PORTFOLIO_NAMES.get(risk_level, PORTFOLIO_NAMES['平衡型'])
    result = {}
    if style:
        pool = names_dict.get(style, [])
        selected = random.sample(pool, min(3, len(pool)))
        result[style] = selected
    else:
        for s, pool in names_dict.items():
            result[s] = random.sample(pool, min(3, len(pool)))
    return result


# ══════════════════════════════════════════════════════════════
# 组合 CRUD
# ══════════════════════════════════════════════════════════════

def list_portfolios(user_id):
    rows = _db_execute(
        "SELECT * FROM portfolios WHERE is_active=1 AND user_id=%s ORDER BY updated_at DESC",
        (user_id,), fetch=True)
    if not rows:
        return []
    result = []
    for r in rows:
        metrics = _compute_portfolio_summary(r['id'])
        result.append({
            'id': r['id'],
            'name': r['name'],
            'description': r.get('description', ''),
            'tags': json.loads(r['tags']) if r.get('tags') else [],
            'risk_level': r['risk_level'],
            'target_max_drawdown': float(r['target_max_drawdown']) if r.get('target_max_drawdown') else None,
            'created_from': r.get('created_from', 'custom'),
            'created_at': str(r.get('created_at', '')),
            **metrics,
        })
    return result


def _compute_portfolio_summary(portfolio_id):
    """计算组合摘要指标（轻量版，仅统计数量不计算NAV）"""
    holdings = _db_execute(
        "SELECT * FROM portfolio_holdings WHERE portfolio_id=%s", (portfolio_id,), fetch=True)
    if not holdings:
        return {'fund_count': 0, 'total_return': None, 'annual_return': None, 'max_drawdown_1y': None}
    return {
        'fund_count': len(holdings),
        'total_return': None,
        'annual_return': None,
        'max_drawdown_1y': None,
    }


def create_portfolio(data, user_id):
    name = data.get('name', '')
    mode = data.get('mode', 'custom')
    risk_level = data.get('risk_level', '平衡型')
    description = data.get('description', '')
    tags = json.dumps(data.get('tags', []), ensure_ascii=False)
    target_dd = data.get('target_max_drawdown')
    fund_codes = data.get('fund_codes', [])
    fund_weights = data.get('fund_weights', [])
    pid = _db_execute(
        "INSERT INTO portfolios (user_id, name, description, tags, risk_level, target_max_drawdown, created_from) "
        "VALUES (%s, %s, %s, %s, %s, %s, %s)",
        (user_id, name, description, tags, risk_level, target_dd, mode), fetch=True)
    if not isinstance(pid, int):
        pid = _db_execute("SELECT LAST_INSERT_ID() as id", fetch=True)
        pid = pid[0]['id'] if pid else 1
    if fund_codes and fund_weights:
        _add_funds_to_portfolio(pid, fund_codes, fund_weights)
    return {'success': True, 'portfolio_id': pid, 'name': name}


def get_portfolio(portfolio_id, user_id=None):
    if user_id:
        rows = _db_execute(
            "SELECT * FROM portfolios WHERE id=%s AND is_active=1 AND user_id=%s",
            (portfolio_id, user_id), fetch=True)
    else:
        rows = _db_execute(
            "SELECT * FROM portfolios WHERE id=%s AND is_active=1", (portfolio_id,), fetch=True)
    if not rows:
        return None
    r = rows[0]
    holdings = get_holdings(portfolio_id)
    return {
        'id': r['id'],
        'name': r['name'],
        'description': r.get('description', ''),
        'tags': json.loads(r['tags']) if r.get('tags') else [],
        'risk_level': r['risk_level'],
        'target_max_drawdown': float(r['target_max_drawdown']) if r.get('target_max_drawdown') else None,
        'created_from': r.get('created_from', 'custom'),
        'created_at': str(r.get('created_at', '')),
        'updated_at': str(r.get('updated_at', '')),
        'holdings': holdings,
    }


def update_portfolio(portfolio_id, data, user_id=None):
    # Verify ownership
    portfolio = get_portfolio(portfolio_id, user_id)
    if portfolio is None:
        return {'success': False, 'error': '组合不存在或无权访问'}
    fields = []
    values = []
    allowed = ['name', 'description', 'tags', 'risk_level', 'target_max_drawdown']
    for key in allowed:
        if key in data:
            fields.append(f"{key}=%s")
            values.append(json.dumps(data[key], ensure_ascii=False) if key == 'tags' else data[key])
    if not fields:
        return {'success': False, 'error': '没有可更新的字段'}
    fields.append("updated_at=%s")
    values.append(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    values.append(portfolio_id)
    _db_execute(f"UPDATE portfolios SET {', '.join(fields)} WHERE id=%s", values, fetch=False)
    return {'success': True}


def delete_portfolio(portfolio_id, user_id=None):
    # Verify ownership
    portfolio = get_portfolio(portfolio_id, user_id)
    if portfolio is None:
        return {'success': False, 'error': '组合不存在或无权访问'}
    _db_execute("DELETE FROM portfolio_holdings WHERE portfolio_id=%s", (portfolio_id,), fetch=False)
    _db_execute("DELETE FROM portfolio_nav_cache WHERE portfolio_id=%s", (portfolio_id,), fetch=False)
    _db_execute("DELETE FROM portfolio_backtest_cache WHERE portfolio_id=%s", (portfolio_id,), fetch=False)
    _db_execute("UPDATE portfolios SET is_active=0 WHERE id=%s", (portfolio_id,), fetch=False)
    return {'success': True}


def duplicate_portfolio(portfolio_id, user_id=None):
    # Verify ownership
    portfolio = get_portfolio(portfolio_id, user_id)
    if portfolio is None:
        return {'success': False, 'error': '组合不存在或无权访问'}
    original = _db_execute(
        "SELECT * FROM portfolios WHERE id=%s", (portfolio_id,), fetch=True)
    if not original:
        return {'success': False, 'error': '组合不存在'}
    o = original[0]
    new_name = f"{o['name']}（副本）"
    new_pid = _db_execute(
        "INSERT INTO portfolios (user_id, name, description, tags, risk_level, target_max_drawdown, created_from) "
        "VALUES (%s, %s, %s, %s, %s, %s, 'custom')",
        (user_id, new_name, o.get('description', ''),
         o.get('tags', '[]'), o['risk_level'],
         o.get('target_max_drawdown')), fetch=True)
    if not isinstance(new_pid, int):
        r = _db_execute("SELECT LAST_INSERT_ID() as id", fetch=True)
        new_pid = r[0]['id'] if r else 1
    holdings = _db_execute(
        "SELECT * FROM portfolio_holdings WHERE portfolio_id=%s", (portfolio_id,), fetch=True)
    if holdings:
        for h in holdings:
            _db_execute(
                "INSERT INTO portfolio_holdings (portfolio_id, fund_code, fund_name, fund_type, weight) "
                "VALUES (%s, %s, %s, %s, %s)",
                (new_pid, h['fund_code'], h.get('fund_name', ''), h.get('fund_type', ''), h['weight']),
                fetch=False)
    return {'success': True, 'portfolio_id': new_pid, 'name': new_name}


# ══════════════════════════════════════════════════════════════
# 持仓管理
# ══════════════════════════════════════════════════════════════

def get_holdings(portfolio_id):
    rows = _db_execute(
        "SELECT * FROM portfolio_holdings WHERE portfolio_id=%s ORDER BY weight DESC",
        (portfolio_id,), fetch=True)
    if not rows:
        return []
    return [{
        'id': r['id'], 'fund_code': r['fund_code'],
        'fund_name': r.get('fund_name', ''), 'fund_type': r.get('fund_type', ''),
        'weight': float(r['weight']),
        'added_at': str(r.get('added_at', '')),
    } for r in rows]


def _add_funds_to_portfolio(portfolio_id, fund_codes, weights):
    """向组合添加基金"""
    for fc, w in zip(fund_codes, weights):
        fund_name = _resolve_fund_name(fc)
        fund_type_str = _guess_fund_type_by_name(fund_name)
        try:
            _db_execute(
                "INSERT INTO portfolio_holdings (portfolio_id, fund_code, fund_name, fund_type, weight) "
                "VALUES (%s, %s, %s, %s, %s)",
                (portfolio_id, fc, fund_name, fund_type_str, w), fetch=False)
        except Exception:
            _db_execute(
                "UPDATE portfolio_holdings SET weight=%s WHERE portfolio_id=%s AND fund_code=%s",
                (w, portfolio_id, fc), fetch=False)
    _invalidate_nav_cache(portfolio_id)


def add_holdings(portfolio_id, fund_codes, weights=None):
    if not fund_codes:
        return {'success': False, 'error': '请选择至少一只基金'}
    if len(fund_codes) > 10:
        return {'success': False, 'error': '一次最多添加10只基金'}
    existing = get_holdings(portfolio_id)
    if len(existing) + len(fund_codes) > 15:
        return {'success': False, 'error': '组合最多持有15只基金'}
    if weights is None:
        weights = [20.0 / len(fund_codes)] * len(fund_codes)
    for w in weights:
        if w > 20:
            return {'success': False, 'error': '单只基金权重不能超过20%'}
    _add_funds_to_portfolio(portfolio_id, fund_codes, weights)
    return {'success': True, 'message': f'已添加 {len(fund_codes)} 只基金'}


def update_holding(holding_id, weight):
    if weight <= 0 or weight > 20:
        return {'success': False, 'error': '权重需在0-20%之间'}
    _db_execute("UPDATE portfolio_holdings SET weight=%s WHERE id=%s", (weight, holding_id), fetch=False)
    row = _db_execute("SELECT portfolio_id FROM portfolio_holdings WHERE id=%s", (holding_id,), fetch=True)
    if row:
        _invalidate_nav_cache(row[0]['portfolio_id'])
    return {'success': True}


def remove_holding(holding_id):
    row = _db_execute("SELECT portfolio_id FROM portfolio_holdings WHERE id=%s", (holding_id,), fetch=True)
    _db_execute("DELETE FROM portfolio_holdings WHERE id=%s", (holding_id,), fetch=False)
    if row:
        _invalidate_nav_cache(row[0]['portfolio_id'])
    return {'success': True}


def _resolve_fund_name(fund_code):
    """解析基金名称 - 优先从 fund_list_cache 查找"""
    try:
        rows = _db_execute(
            "SELECT name FROM fund_list_cache WHERE code=%s LIMIT 1",
            (fund_code,), fetch=True)
        if rows and len(rows) > 0 and rows[0].get('name'):
            return rows[0]['name']
    except Exception:
        pass
    try:
        _ensure_deps()
        import app
        name = app.get_fund_name(fund_code)
        if name and name != fund_code:
            return name
    except Exception:
        pass
    fund_data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'fund_data')
    try:
        fpath = os.path.join(fund_data_dir, f'{fund_code}.json')
        if os.path.exists(fpath):
            with open(fpath, 'r') as f:
                data = json.load(f)
            if data.get('基金简称'):
                return data['基金简称']
    except Exception:
        pass
    return fund_code


def _guess_fund_type_by_name(name):
    """根据基金名称推断类型（增强版）"""
    if not name:
        return '混合型'
    n = str(name)
    if any(kw in n for kw in ['货币', '现金', '货基']):
        return '货币型'
    if any(kw in n for kw in ['债券', '纯债', '短债', '中短债', '信用债', '利率债', '可转债', '国债']):
        return '债券型'
    if any(kw in n for kw in ['ETF', 'etf', '指数']):
        return '指数型'
    if any(kw in n for kw in ['股票', '成长']):
        return '股票型'
    if any(kw in n for kw in ['QDII', 'qdii', '全球', '海外']):
        return 'QDII型'
    if any(kw in n for kw in ['混合', '灵活', '平衡', '配置', '稳健', '优选', '精选', '蓝筹', '红利', '消费', '医疗',
                               '医药', '科技', '新能源', '制造', '主题', '轮动', '量化', '多因子', '驱动',
                               '创新', '升级', '转型', '改革', '动力', '趋势', '领先', '龙头']):
        return '混合型'
    return '混合型'


_FUND_POOL_CACHE = None
_FUND_POOL_CACHE_TIME = 0

def _build_fund_pool(force_refresh=False):
    """构建全市场基金池：合并 fund_list_cache（名称）+ fund_basic（指标）+ fund_data JSON"""
    global _FUND_POOL_CACHE, _FUND_POOL_CACHE_TIME
    import time as _t
    if not force_refresh and _FUND_POOL_CACHE is not None and _t.time() - _FUND_POOL_CACHE_TIME < 3600:
        return _FUND_POOL_CACHE.copy()
    pool = {}
    # 1. 从 SQLite fund_list_cache 获取所有基金名称（限制5000只以提高性能）
    try:
        rows = _db_execute("SELECT code, name FROM fund_list_cache LIMIT 5000", fetch=True)
        if rows:
            for r in rows:
                code = r.get('code', '')
                name = r.get('name', '')
                if code and name:
                    pool[code] = {
                        'code': code, 'name': name, 'type': _guess_fund_type_by_name(name),
                        'annual_return': 0, 'max_drawdown': 0, 'sharpe': 0, 'manager': '', 'source': '列表'
                    }
    except Exception as e:
        print(f"[portfolio] fund_list_cache read failed: {e}")
    # 2. 覆盖/补充 fund_basic 的指标数据
    try:
        basic_rows = _db_execute(
            "SELECT fund_code, fund_name, fund_style, fund_manager, annual_return, max_drawdown, "
            "sharpe_ratio, annual_volatility FROM fund_basic", fetch=True)
        if basic_rows:
            for r in basic_rows:
                fc = r.get('fund_code', '')
                if not fc:
                    continue
                if fc not in pool:
                    pool[fc] = {'code': fc, 'name': r.get('fund_name') or fc, 'type': '混合型',
                                'annual_return': 0, 'max_drawdown': 0, 'sharpe': 0, 'manager': '', 'source': '基础'}
                # 更新名称（fund_basic的名称可能缺失）
                if r.get('fund_name'):
                    pool[fc]['name'] = r['fund_name']
                # 更新类型根据风格
                st = r.get('fund_style', '') or ''
                if '债券' in str(st):
                    pool[fc]['type'] = '债券型'
                elif '指数' in str(st):
                    pool[fc]['type'] = '指数型'
                try:
                    pool[fc]['annual_return'] = float(str(r.get('annual_return', '0')).replace('%', ''))
                except Exception:
                    pass
                try:
                    pool[fc]['max_drawdown'] = float(str(r.get('max_drawdown', '0')).replace('%', ''))
                except Exception:
                    pass
                try:
                    pool[fc]['sharpe'] = float(r.get('sharpe_ratio', 0) or 0)
                except Exception:
                    pass
                pool[fc]['manager'] = str(r.get('fund_manager', '') or '')
                pool[fc]['source'] = '基础'
    except Exception as e:
        print(f"[portfolio] fund_basic read failed: {e}")
    # 3. 用 fund_data JSON 覆盖/补充
    fund_data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'fund_data')
    try:
        for fname in os.listdir(fund_data_dir):
            if not fname.endswith('.json') or fname == 'load_time.json':
                continue
            fc = fname.replace('.json', '')
            try:
                with open(os.path.join(fund_data_dir, fname), 'r') as f:
                    data = json.load(f)
                if fc not in pool:
                    pool[fc] = {'code': fc, 'name': data.get('基金简称', fc), 'type': '混合型',
                                'annual_return': 0, 'max_drawdown': 0, 'sharpe': 0, 'manager': '', 'source': '本地'}
                if data.get('基金简称'):
                    pool[fc]['name'] = data['基金简称']
                try:
                    pool[fc]['annual_return'] = float(str(data.get('年化收益率', '0')).replace('%', ''))
                except Exception:
                    pass
                try:
                    pool[fc]['max_drawdown'] = float(str(data.get('最大回撤', '0')).replace('%', ''))
                except Exception:
                    pass
                try:
                    pool[fc]['sharpe'] = float(data.get('夏普比率', 0) or 0)
                except Exception:
                    pass
                pool[fc]['manager'] = data.get('基金经理', '') or ''
                pool[fc]['source'] = '本地'
            except Exception:
                continue
    except Exception:
        pass
    _FUND_POOL_CACHE = pool
    _FUND_POOL_CACHE_TIME = _t.time()
    print(f"[portfolio] Fund pool built: {len(pool)} funds")
    return pool.copy()


def _score_fund_for_type(f, target_type):
    """为特定资产类型评分一只基金（越高越好）"""
    score = 50
    ftype = f.get('type', '混合型')
    # 类型精确匹配加分
    if ftype == target_type:
        score += 25
    # 指标加分（过滤明显异常的数据）
    ar = f.get('annual_return', 0) or 0
    sharpe = f.get('sharpe', 0) or 0
    dd = f.get('max_drawdown', 0) or 0
    # 过滤异常值：年化>60%或<-30%很可能是数据问题
    if ar > 60 or ar < -30:
        score -= 40
    if ar > 3 and ar < 60:
        score += min(ar * 0.5, 25)
    if sharpe > 0.1 and sharpe < 8:
        score += min(sharpe * 10, 25)
    # 回撤越小得分越高：-5%回撤得12分，-70%回撤得2分
    if dd < 0 and dd > -80:
        score += max(0, 12 - abs(dd) / 7)
    # 经理加分
    manager = f.get('manager', '')
    if manager and len(str(manager)) > 3:
        score += 6
    return score


# 各风险等级允许的单只基金最大回撤上限 & 组合整体回撤目标
RISK_MAX_DRAWDOWN = {
    '保守型': -5,    # 单只≤5%，组合目标≤3%
    '稳健型': -10,   # 单只≤10%，组合目标≤7%
    '平衡型': -15,   # 单只≤15%，组合目标≤12%
    '成长型': -25,   # 单只≤25%，组合目标≤20%
    '激进型': -35,   # 单只≤35%，组合目标≤30%
}
# 组合整体回撤目标（与PRD风险等级对应，考虑分散化效应后略低于单只上限）
PORTFOLIO_DD_TARGET = {
    '保守型': 5.0,   # PRD: 保守型≤5%
    '稳健型': 10.0,  # PRD: 稳健型5-10%
    '平衡型': 15.0,  # PRD: 平衡型10-15%
    '成长型': 25.0,  # PRD: 成长型15-25%
    '激进型': 35.0,  # PRD: 激进型25-35%
}
# 资产类型的预估回撤（用于无数据的基金）
ASSET_TYPE_TYPICAL_DD = {
    '货币型': -0.5,
    '债券型': -3.0,
    '混合型': -15.0,
    '股票型': -25.0,
    '指数型': -22.0,
    'QDII型': -20.0,
}


def _dedup_share_classes(candidates):
    """同一基金的不同份额(A/C/D/E等)只保留评分最高的那只"""
    import re
    seen_base = {}
    result = []
    for score, f in candidates:
        name = f.get('name', '')
        # 去除份额后缀(A/B/C/D/E等单字母)得到基础名称
        base_name = re.sub(r'[A-EH]$', '', name.strip())
        # 也尝试按代码前3位分组（更激进）
        code = f.get('code', '')
        code_prefix = code[:3] if len(code) >= 3 else code
        # 使用基础名称作为去重键
        key = base_name if base_name else code_prefix
        if key not in seen_base or score > seen_base[key][0]:
            seen_base[key] = (score, f)
    return sorted(seen_base.values(), key=lambda x: x[0], reverse=True)


def _pick_funds_for_type(pool, target_type, target_pct, used_codes, risk_level='平衡型', max_per_fund=20, max_funds=3):
    """为特定资产类型从基金池中挑选最优基金（排除已选，回撤过滤，同基金去重）"""
    max_dd = RISK_MAX_DRAWDOWN.get(risk_level, -15)
    candidates = []
    for fc, f in pool.items():
        if fc in used_codes:
            continue
        ftype = f.get('type', '混合型')
        if ftype != target_type:
            continue
        # 回撤过滤：单只基金回撤必须≤风险等级上限
        dd = f.get('max_drawdown', 0) or 0
        # 无回撤数据的默认允许（但给予低分）
        if dd < 0 and dd < max_dd:
            continue  # 回撤超过该风险等级上限，排除
        # 过滤明显异常的年化收益数据(>50%几乎都是数据错误)
        ar = f.get('annual_return', 0) or 0
        if ar > 50 or ar < -20:
            continue
        s = _score_fund_for_type(f, target_type)
        candidates.append((s, f))
    # 按评分排序
    candidates.sort(key=lambda x: x[0], reverse=True)
    # 同基金不同份额去重
    candidates = _dedup_share_classes(candidates)
    selected = []
    remaining = target_pct
    if not candidates:
        return selected, remaining
    top = candidates[:max_funds]
    if not top:
        return selected, remaining
    per_fund = min(target_pct / len(top), max_per_fund)
    for score, f in top:
        if remaining <= 0.5:
            break
        w = min(per_fund, remaining, max_per_fund)
        if w < 1:
            break
        selected.append({
            'fund_code': f['code'],
            'fund_name': f.get('name', f['code']),
            'fund_type': target_type,
            'weight': round(w, 1),
            'score': round(score, 0),
            'annual_return': f.get('annual_return', 0),
            'max_drawdown': f.get('max_drawdown', 0),
            'sharpe': f.get('sharpe', 0),
            'fund_manager': f.get('manager', ''),
            'reason': _build_reason(f, target_type),
        })
        remaining -= w
    return selected, remaining


def _build_reason(fund, ftype):
    """构建详细的基金选择理由"""
    ar = fund.get('annual_return', 0) or 0
    sharpe = fund.get('sharpe', 0) or 0
    dd = fund.get('max_drawdown', 0) or 0
    manager = fund.get('manager', '')
    name = fund.get('name', '')
    parts = []
    # 基金特征
    if '债券' in str(name):
        parts.append('纯债/债券型，波动低收益稳')
    elif '混合' in str(name) or '灵活' in str(name):
        parts.append('混合型，股债灵活配置')
    elif '股票' in str(name):
        parts.append('股票型，专注权益投资')
    elif '指数' in str(name) or 'ETF' in str(name):
        parts.append('指数型，跟踪市场基准')
    elif '货币' in str(name):
        parts.append('货币型，流动性好')
    # 量化指标
    if ar > 3 and ar < 60:
        parts.append(f'年化回报{ar:.1f}%')
    if sharpe > 0.3 and sharpe < 8:
        parts.append(f'夏普比率{sharpe:.2f}(风险调整收益良好)')
    if dd < 0 and dd > -80:
        parts.append(f'最大回撤{dd:.1f}%')
    # 基金经理
    if manager and len(str(manager)) > 3:
        mgr_name = str(manager).split('(')[0].strip().split(' ')[0][:6]
        if mgr_name:
            parts.append(f'基金经理{mgr_name}')
    if not parts:
        parts.append(f'{ftype}标的，资产配置需要')
    return '；'.join(parts)


# ══════════════════════════════════════════════════════════════
# 净值计算与缓存
# ══════════════════════════════════════════════════════════════

def _get_cached_portfolio_nav(portfolio_id):
    rows = _db_execute(
        "SELECT nav_date, nav_value FROM portfolio_nav_cache WHERE portfolio_id=%s "
        "ORDER BY nav_date ASC", (portfolio_id,), fetch=True)
    if not rows:
        return None
    return [{'nav_date': str(r['nav_date']), 'nav_value': float(r['nav_value'])} for r in rows]


def _cache_portfolio_nav(portfolio_id, nav_df):
    """将组合净值批量缓存到数据库（MySQL 优先，SQLite 降级）"""
    rows = []
    for _, row in nav_df.iterrows():
        d = str(row.get('nav_date', ''))[:10]
        if not d:
            continue
        rows.append((portfolio_id, d, float(row['nav'])))
    if not rows:
        return
    # 批量插入：每次 200 条，构建多行 VALUES
    batch_size = 200
    for i in range(0, len(rows), batch_size):
        batch = rows[i:i + batch_size]
        placeholders = ', '.join(['(%s, %s, %s)'] * len(batch))
        flat_params = [v for row in batch for v in row]
        try:
            _db_execute(
                f"INSERT IGNORE INTO portfolio_nav_cache (portfolio_id, nav_date, nav_value) "
                f"VALUES {placeholders}",
                tuple(flat_params), fetch=False)
        except Exception:
            try:
                q_placeholders = ', '.join(['(?, ?, ?)'] * len(batch))
                _db_execute(
                    f"INSERT OR IGNORE INTO portfolio_nav_cache (portfolio_id, nav_date, nav_value) "
                    f"VALUES {q_placeholders}",
                    tuple(flat_params), fetch=False)
            except Exception:
                pass


def _invalidate_nav_cache(portfolio_id):
    try:
        _db_execute("DELETE FROM portfolio_nav_cache WHERE portfolio_id=%s", (portfolio_id,), fetch=False)
    except Exception:
        pass
    try:
        _db_execute("DELETE FROM portfolio_backtest_cache WHERE portfolio_id=%s", (portfolio_id,), fetch=False)
    except Exception:
        pass


def compute_portfolio_nav_data(portfolio_id, years=3):
    """计算组合加权净值序列（并行获取各基金 NAV，大幅减少多基金组合耗时）"""
    holdings = get_holdings(portfolio_id)
    if not holdings:
        return None, "组合中没有基金"
    # 检查缓存（>=2 条即可用，避免重复计算）
    cached = _get_cached_portfolio_nav(portfolio_id)
    if cached and len(cached) >= 2:
        return cached, None
    # 并行获取每只基金的净值数据
    fund_navs = {}
    try:
        from fund_crawler import crawl_fund_nav_df
    except Exception:
        return None, "无法导入数据模块"

    def _fetch_one_fund_nav(fund_code):
        """获取单只基金 NAV 并转为 Series"""
        try:
            nav_records = crawl_fund_nav_df(fund_code, years=years)
            if nav_records and len(nav_records) > 0:
                records = []
                for rec in nav_records:
                    if isinstance(rec, dict):
                        records.append({'date': str(rec.get('date', rec.get('净值日期', ''))),
                                        'nav': float(rec.get('nav', rec.get('单位净值', rec.get('net_value', 0))))})
                    elif isinstance(rec, (list, tuple)) and len(rec) >= 2:
                        records.append({'date': str(rec[0]), 'nav': float(rec[1])})
                if records:
                    df = pd.DataFrame(records)
                    df['date'] = pd.to_datetime(df['date'])
                    df = df.set_index('date').sort_index()
                    return fund_code, df['nav']
        except Exception as e:
            print(f"[portfolio] Failed to get NAV for {fund_code}: {e}")
        return fund_code, None

    # 基金数量少时直接串行（避免线程开销），多时并行
    if len(holdings) <= 3:
        for h in holdings:
            _fc, _nav = _fetch_one_fund_nav(h['fund_code'])
            if _nav is not None:
                fund_navs[_fc] = _nav
    else:
        with concurrent.futures.ThreadPoolExecutor(max_workers=min(len(holdings), 10)) as executor:
            futures = {executor.submit(_fetch_one_fund_nav, h['fund_code']): h['fund_code'] for h in holdings}
            for future in concurrent.futures.as_completed(futures):
                try:
                    _fc, _nav = future.result()
                    if _nav is not None:
                        fund_navs[_fc] = _nav
                except Exception as e:
                    print(f"[portfolio] NAV fetch task failed: {e}")

    if not fund_navs:
        return None, "无法获取基金净值数据"
    # 对齐日期并计算加权净值
    all_dates = set()
    for nav in fund_navs.values():
        all_dates.update(nav.index)
    all_dates = sorted(all_dates)
    if len(all_dates) < 2:
        return None, "数据不足"
    portfolio_nav = []
    for d in all_dates:
        weighted_nav = 0
        total_weight = 0
        for h in holdings:
            fc = h['fund_code']
            w = float(h['weight']) / 100
            if fc in fund_navs:
                nav_series = fund_navs[fc]
                closest_date = min(nav_series.index, key=lambda x: abs((x - d).days) if hasattr(x, 'days') else abs((pd.Timestamp(x) - pd.Timestamp(d)).days))
                if abs((closest_date - d).days) <= 5:
                    base_nav = nav_series.iloc[0]
                    current_nav = nav_series.loc[closest_date]
                    if base_nav > 0:
                        normalized = current_nav / base_nav
                        weighted_nav += w * normalized
                        total_weight += w
        if total_weight > 0:
            portfolio_nav.append({
                'nav_date': str(d)[:10] if hasattr(d, 'strftime') else str(pd.Timestamp(d))[:10],
                'nav_value': round(weighted_nav / total_weight, 4),
            })
    if portfolio_nav:
        nav_df = pd.DataFrame([{'nav_date': r['nav_date'], 'nav': r['nav_value']} for r in portfolio_nav])
        try:
            _cache_portfolio_nav(portfolio_id, nav_df)
        except Exception:
            pass
    return portfolio_nav, None


def _compute_portfolio_metrics(portfolio_id):
    """计算组合核心指标"""
    nav_data, _ = compute_portfolio_nav_data(portfolio_id, years=3)
    if not nav_data or len(nav_data) < 2:
        return {}
    navs = [d['nav_value'] for d in nav_data]
    nav_series = pd.Series(navs)
    total_return = (nav_series.iloc[-1] / nav_series.iloc[0] - 1) * 100
    days = len(nav_series)
    years = days / 252
    annual_return = ((1 + total_return / 100) ** (1 / max(years, 0.05)) - 1) * 100
    daily_returns = nav_series.pct_change().dropna()
    annual_volatility = daily_returns.std() * np.sqrt(252) * 100 if len(daily_returns) > 0 else 0
    rf = 2.5
    sharpe = (annual_return - rf) / annual_volatility if annual_volatility > 0 else 0
    downside = daily_returns[daily_returns < 0]
    downside_vol = downside.std() * np.sqrt(252) * 100 if len(downside) > 0 else annual_volatility
    sortino = (annual_return - rf) / downside_vol if downside_vol > 0 else 0
    peak = nav_series.expanding().max()
    drawdown = (nav_series / peak - 1) * 100
    max_dd = float(drawdown.min())
    positive_days = int((daily_returns > 0).sum())
    win_rate = (positive_days / len(daily_returns) * 100) if len(daily_returns) > 0 else 0
    return {
        'total_return': round(total_return, 2),
        'annual_return': round(annual_return, 2),
        'annual_volatility': round(annual_volatility, 2),
        'sharpe_ratio': round(sharpe, 2),
        'sortino_ratio': round(sortino, 2),
        'max_drawdown': round(max_dd, 2),
        'win_rate': round(win_rate, 1),
        'data_days': days,
    }


# ══════════════════════════════════════════════════════════════
# 组合分析（行业/资产/公司分布）
# ══════════════════════════════════════════════════════════════

def compute_portfolio_analysis(portfolio_id):
    holdings = get_holdings(portfolio_id)
    if not holdings:
        return {'asset_allocation': [], 'sector_allocation': [], 'company_allocation': []}
    asset_map = {}
    for h in holdings:
        ft = h.get('fund_type', '混合型')
        asset_map[ft] = asset_map.get(ft, 0) + float(h['weight'])
    asset_allocation = [{'type': k, 'weight': round(v, 1)} for k, v in
                       sorted(asset_map.items(), key=lambda x: x[1], reverse=True)]
    return {
        'asset_allocation': asset_allocation,
        'sector_allocation': [],
        'company_allocation': [],
        'holdings_analysis': _build_holdings_analysis(holdings),
    }


def _build_holdings_analysis(holdings):
    """构建持仓分析数据，优先从缓存获取日增长率"""
    result = []
    for h in holdings:
        fc = h['fund_code']
        day_growth = ''
        fund_name = h.get('fund_name', fc)
        # 1. 尝试从Redis/内存缓存获取fund info
        try:
            _ensure_deps()
            ck = _generate_cache_key('fund:info', fc)
            info = _get_cache(ck)
            if info:
                day_growth = str(info.get('日增长率', ''))
                if info.get('基金简称'):
                    fund_name = info['基金简称']
        except Exception:
            pass
        # 2. 尝试从SQLite fund_basic获取
        if not day_growth:
            try:
                rows = _db_execute(
                    "SELECT day_growth FROM fund_basic WHERE fund_code=%s LIMIT 1",
                    (fc,), fetch=True)
                if rows and len(rows) > 0:
                    dg = rows[0].get('day_growth', '')
                    if dg:
                        day_growth = str(dg)
            except Exception:
                pass
        # 3. 尝试从JSON文件获取
        if not day_growth:
            try:
                fund_data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'fund_data')
                fpath = os.path.join(fund_data_dir, f'{fc}.json')
                if os.path.exists(fpath):
                    with open(fpath, 'r') as fp:
                        jd = json.load(fp)
                    dg = jd.get('日增长率', '')
                    if dg:
                        day_growth = str(dg)
            except Exception:
                pass
        result.append({
            'fund_code': fc,
            'fund_name': fund_name,
            'fund_type': h.get('fund_type', ''),
            'weight': float(h['weight']),
            'day_growth': day_growth,
        })
    return result


# ══════════════════════════════════════════════════════════════
# 基金搜索与建议
# ══════════════════════════════════════════════════════════════

def search_funds_for_portfolio(keyword):
    """增强搜索：按代码/名称返回基金列表，附带简单指标"""
    _ensure_deps()
    if not keyword or len(keyword) < 1:
        return []
    try:
        import app
        from fund_crawler import crawl_fund_full
    except Exception:
        return []
    results = []
    # 先精确代码搜索
    if len(keyword) == 6 and keyword.isdigit():
        try:
            fund_name = app.get_fund_name(keyword)
            results.append({
                'fund_code': keyword, 'fund_name': fund_name,
                'fund_type': _guess_fund_type_by_name(fund_name),
                'source': '代码匹配',
            })
        except Exception:
            pass
    # 模糊搜索
    try:
        search_results = app.search_fund()
        if hasattr(search_results, 'get_json'):
            pass
    except Exception:
        pass
    # 从 SQLite fund_list_cache 搜索
    try:
        rows = _db_execute(
            "SELECT code, name FROM fund_list_cache WHERE code LIKE %s OR name LIKE %s LIMIT 15",
            (f'%{keyword}%', f'%{keyword}%'), fetch=True)
        if rows:
            for r in rows:
                if len(results) >= 15:
                    break
                exists = any(x['fund_code'] == r['code'] for x in results)
                if not exists:
                    results.append({
                        'fund_code': r['code'], 'fund_name': r['name'],
                        'fund_type': _guess_fund_type_by_name(r['name']),
                        'source': '基金库',
                    })
    except Exception:
        pass
    # 从常用基金JSON文件搜索
    if len(results) < 5:
        fund_data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'fund_data')
        try:
            for fname in os.listdir(fund_data_dir):
                if fname.endswith('.json') and fname != 'load_time.json':
                    code = fname.replace('.json', '')
                    if keyword in code:
                        exists = any(x['fund_code'] == code for x in results)
                        if not exists:
                            try:
                                with open(os.path.join(fund_data_dir, fname), 'r') as f:
                                    data = json.load(f)
                                    results.append({
                                        'fund_code': code,
                                        'fund_name': data.get('基金简称', code),
                                        'fund_type': _guess_fund_type_by_name(data.get('基金简称', '')),
                                        'source': '本地数据',
                                    })
                            except Exception:
                                pass
        except Exception:
            pass
    return results[:15]


def get_suggested_funds(risk_type=None):
    """获取系统建议基金（按风险类型分类），从全市场基金池筛选"""
    _ensure_deps()
    cache_key = _generate_cache_key('portfolio', 'suggested_funds_v2')
    cached = _get_cache(cache_key)
    if cached:
        if risk_type:
            return cached.get(risk_type, [])
        return cached
    pool = _build_fund_pool()
    # 按风险类型分类：低风险(保守/稳健)选债券+货币，中风险(平衡)选混合，高风险(成长/激进)选股票+指数
    suggested = {'保守型': [], '稳健型': [], '平衡型': [], '成长型': [], '激进型': []}
    for fc, f in pool.items():
        ftype = f.get('type', '混合型')
        entry = {
            'fund_code': fc, 'fund_name': f.get('name', fc), 'fund_type': ftype,
            'annual_return': f.get('annual_return', 0),
            'max_drawdown': f.get('max_drawdown', 0),
            'sharpe_ratio': f.get('sharpe', 0),
            'fund_manager': f.get('manager', ''),
        }
        # 保守型：债券+货币
        if ftype in ('债券型', '货币型'):
            suggested['保守型'].append(entry)
            suggested['稳健型'].append(entry)
        # 稳健型：+混合型(低回撤)
        if ftype == '混合型' and (f.get('max_drawdown', 0) or 0) > -30:
            suggested['稳健型'].append(entry)
            suggested['平衡型'].append(entry)
        # 平衡型：混合型(全部)
        if ftype == '混合型':
            suggested['平衡型'].append(entry)
            suggested['成长型'].append(entry)
        # 成长型：股票+指数
        if ftype in ('股票型', '指数型'):
            suggested['成长型'].append(entry)
            suggested['激进型'].append(entry)
        # 激进型：所有类型
        suggested['激进型'].append(entry)
    # 每个分类按综合评分排序取前 20
    for risk in suggested:
        suggested[risk].sort(key=lambda x: (x['annual_return'] or 0) * 2 + (x['sharpe_ratio'] or 0) * 10 - abs(x['max_drawdown'] or 0) * 0.5, reverse=True)
        suggested[risk] = suggested[risk][:20]
    _set_cache(cache_key, suggested, 3600)
    if risk_type:
        return suggested.get(risk_type, [])
    return suggested


# ══════════════════════════════════════════════════════════════
# 组合推荐
# ══════════════════════════════════════════════════════════════

def generate_recommendation(risk_level='平衡型'):
    """根据风险等级，从全市场基金池中按资产配置比例生成推荐组合"""
    try:
        return _generate_recommendation_impl(risk_level)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {
            'risk_level': risk_level,
            'core_idea': '推荐生成失败，请稍后重试',
            'asset_allocation': ASSET_ALLOCATION.get(risk_level, {}),
            'funds': [],
            'fund_count': 0,
            'total_weight': 0,
            'suggested_hold_period': '1年以上',
            'methodology': '',
            'risk_warnings': [],
            'error': str(e),
        }


def _asset_typical_dd(ftype):
    """获取某类基金的典型最大回撤估算值"""
    return ASSET_TYPE_TYPICAL_DD.get(ftype, -15.0)


def _estimate_portfolio_drawdown(funds):
    """估算组合整体最大回撤（加权平均 + 分散化折扣）

    算法：
    1. 计算加权平均回撤（权重×单只回撤）
    2. 分散化折扣：资产类型越多、低相关资产越多，折扣越大
    3. 组合回撤 ≈ 加权平均 × 分散系数
       - 分散系数 = 1 - 0.05 × (不同类型数 - 1) - 0.03 × (债券/货币占比/10)
    """
    if not funds:
        return 0
    total_w = sum(f['weight'] for f in funds)
    if total_w <= 0:
        return 0
    # 加权平均回撤
    weighted_dd = 0
    for f in funds:
        dd = f.get('max_drawdown', 0) or 0
        if dd >= 0:
            dd = _asset_typical_dd(f.get('fund_type', '混合型'))
        weighted_dd += abs(dd) * f['weight'] / total_w
    # 分散化折扣
    types = set(f.get('fund_type', '混合型') for f in funds)
    bond_weight = sum(f['weight'] for f in funds if f.get('fund_type') == '债券型') / max(total_w, 1)
    money_weight = sum(f['weight'] for f in funds if f.get('fund_type') == '货币型') / max(total_w, 1)
    num_types = len(types)
    # 分散化折扣：资产类型越多、债券/货币占比越高，折扣越大
    # 每种额外资产类型降低6%，债券货币每10%权重降低4%
    discount = 1.0 - 0.06 * (num_types - 1) - 0.04 * (bond_weight + money_weight) / 10
    discount = max(0.3, min(1.0, discount))  # 极限情况保留30%回撤
    portfolio_dd = weighted_dd * discount
    return round(portfolio_dd, 1)


def _generate_recommendation_impl(risk_level):
    """推荐引擎实现"""
    allocation = ASSET_ALLOCATION.get(risk_level, ASSET_ALLOCATION['平衡型'])
    pool = _build_fund_pool()
    selected_funds = []
    used_codes = set()
    remaining_total = 100.0
    # 按目标占比从大到小排序，确保主要类型先选
    sorted_types = sorted(allocation.items(), key=lambda x: x[1], reverse=True)
    for ftype, target_pct in sorted_types:
        if target_pct <= 0 or remaining_total <= 0.5:
            continue
        effective_target = min(target_pct, remaining_total)
        # 根据权重决定选几只（保证总数5-8只）
        if effective_target >= 40:
            max_per_type = 2
        elif effective_target >= 20:
            max_per_type = 2
        else:
            max_per_type = 1
        picked, left = _pick_funds_for_type(pool, ftype, effective_target, used_codes, risk_level=risk_level, max_funds=max_per_type)
        for f in picked:
            used_codes.add(f['fund_code'])
        selected_funds.extend(picked)
        remaining_total -= sum(f['weight'] for f in picked)
    # 若货币型没选到（池子缺少），用债券型补充（需去重+回撤过滤）
    if remaining_total > 5:
        max_dd = RISK_MAX_DRAWDOWN.get(risk_level, -15)
        bond_candidates = [(fc, f) for fc, f in pool.items()
                          if f.get('type') == '债券型' and fc not in used_codes
                          and not ((f.get('max_drawdown', 0) or 0) < 0 and (f.get('max_drawdown', 0) or 0) < max_dd)]
        # 评分排序
        bond_candidates.sort(key=lambda x: _score_fund_for_type(x[1], '债券型'), reverse=True)
        # 同基金不同份额去重
        bond_candidates = _dedup_share_classes(bond_candidates)
        for fc, f in bond_candidates[:2]:
            if remaining_total <= 2:
                break
            w = min(remaining_total, 15.0)
            selected_funds.append({
                'fund_code': fc, 'fund_name': f.get('name', fc), 'fund_type': '债券型',
                'weight': round(w, 1), 'score': 0,
                'annual_return': f.get('annual_return', 0),
                'max_drawdown': f.get('max_drawdown', 0),
                'sharpe': f.get('sharpe', 0),
                'fund_manager': f.get('manager', ''),
                'reason': '补充固收配置，增强组合防御性',
            })
            remaining_total -= w
            used_codes.add(fc)
    # ── 组合整体回撤优化 ──────────────────────────────────────
    target_dd = PORTFOLIO_DD_TARGET.get(risk_level, 12.0)
    for _ in range(8):  # 最多8轮迭代
        est_dd = _estimate_portfolio_drawdown(selected_funds)
        if est_dd <= target_dd + 0.3:
            break  # 达标
        # 按回撤贡献度排序(权重×|回撤|)
        sorted_by_risk = sorted(selected_funds,
            key=lambda f: (f.get('max_drawdown', 0) or _asset_typical_dd(f['fund_type'])) * f['weight'],
            reverse=True)
        sorted_by_safety = sorted(selected_funds,
            key=lambda f: (f.get('max_drawdown', 0) or _asset_typical_dd(f['fund_type'])) * f['weight'])
        # 计算需要降低的回撤量
        excess = est_dd - target_dd
        if excess <= 0:
            break
        # 从最risky向最safe转移：转移量 = min(超出量, risky可转量, safe可增量)
        risky_fund = sorted_by_risk[0]
        safe_fund = sorted_by_safety[0]
        if risky_fund == safe_fund:
            break  # 已经是同一只
        risky_dd = abs(risky_fund.get('max_drawdown', 0) or _asset_typical_dd(risky_fund['fund_type']))
        safe_dd = abs(safe_fund.get('max_drawdown', 0) or _asset_typical_dd(safe_fund['fund_type']))
        # 每转移1%权重减少的回撤 ≈ (risky_dd - safe_dd) / 总权重 / 100
        dd_diff_per_pct = (risky_dd - safe_dd) * 0.7 / 100  # 0.7是分散化折扣
        if dd_diff_per_pct <= 0:
            break
        shift = min(excess / dd_diff_per_pct, risky_fund['weight'] - 5.0, 20.0 - safe_fund['weight'])
        if shift < 0.5:
            # 转移量太小，尝试增加债券型基金
            bond_funds_in_pool = [fc for fc, f in pool.items()
                                 if f.get('type') == '债券型' and fc not in used_codes
                                 and not ((f.get('max_drawdown', 0) or 0) < 0 and (f.get('max_drawdown', 0) or 0) < -10)]
            if bond_funds_in_pool:
                for fc in bond_funds_in_pool[:1]:
                    f = pool[fc]
                    # 减少最risky的权重，增加债券
                    reduce = min(risky_fund['weight'] - 5.0, 10.0)
                    if reduce > 1:
                        risky_fund['weight'] -= reduce
                        selected_funds.append({
                            'fund_code': fc, 'fund_name': f.get('name', fc), 'fund_type': '债券型',
                            'weight': reduce, 'score': 0,
                            'annual_return': f.get('annual_return', 0),
                            'max_drawdown': f.get('max_drawdown', 0),
                            'sharpe': f.get('sharpe', 0),
                            'fund_manager': f.get('manager', ''),
                            'reason': '降低组合回撤，增加固收配置',
                        })
                        used_codes.add(fc)
            break
        risky_fund['weight'] = round(risky_fund['weight'] - shift, 1)
        safe_fund['weight'] = round(safe_fund['weight'] + shift, 1)

    # 最终归一化到100%，同时限制单只≤20%
    for _ in range(3):  # 迭代校正
        total_w = sum(f['weight'] for f in selected_funds)
        if total_w <= 0:
            break
        overages = 0
        for f in selected_funds:
            f['weight'] = round(f['weight'] / total_w * 100, 1)
            if f['weight'] > 20:
                overages += f['weight'] - 20
                f['weight'] = 20.0
        if overages < 0.5:
            break
        # 重新分配超出部分
        under_20 = [f for f in selected_funds if f['weight'] < 20]
        if under_20 and overages > 0:
            add_each = overages / len(under_20)
            for f in under_20:
                f['weight'] = round(f['weight'] + add_each, 1)
    total_w = sum(f['weight'] for f in selected_funds)
    if total_w > 0:
        for f in selected_funds:
            f['weight'] = round(f['weight'] / total_w * 100, 1)
    core_ideas = {
        '保守型': '以本金安全为首要目标，高比例配置债券和货币基金，严格控制回撤，追求稳定增值。基于现代资产组合理论(MPT)，通过低相关性资产的组合最大化风险调整后收益。',
        '稳健型': '稳健为主、适度进取。债券为底仓提供稳定收益，优质混合基金增强回报。在控制最大回撤的前提下，追求超越通胀的长期回报。',
        '平衡型': '股债平衡配置，兼顾进攻与防守。通过资产多元化分散风险，利用不同资产类别的低相关性平滑净值曲线，适应牛熊切换的市场环境。',
        '成长型': '以成长为核心驱动，侧重股票型和指数型基金。通过长期持有优质权益资产获取企业盈利增长带来的超额回报，接受过程中正常的市场波动。',
        '激进型': '高仓位权益配置，积极进取追求高收益。投资于高成长性赛道和优秀基金经理管理的产品，在充分认知风险的前提下，争取最大化长期复利回报。',
    }
    est_dd = _estimate_portfolio_drawdown(selected_funds)
    target_dd = PORTFOLIO_DD_TARGET.get(risk_level, 12.0)
    return {
        'risk_level': risk_level,
        'core_idea': core_ideas.get(risk_level, core_ideas['平衡型']),
        'asset_allocation': allocation,
        'funds': selected_funds,
        'fund_count': len(selected_funds),
        'total_weight': round(sum(f['weight'] for f in selected_funds), 1),
        'estimated_drawdown': est_dd,
        'target_drawdown': target_dd,
        'drawdown_ok': est_dd <= target_dd,
        'suggested_hold_period': {
            '保守型': '6个月以上', '稳健型': '1年以上', '平衡型': '2年以上',
            '成长型': '3年以上', '激进型': '5年以上',
        }.get(risk_level, '1年以上'),
        'methodology': f'基于好买基金4P三性法筛选 + 分散化优化。组合预估最大回撤{est_dd}%（目标≤{target_dd}%），通过{len(set(f.get("fund_type","") for f in selected_funds))}类资产分散配置降低非系统性风险。单只基金权重不超过20%。',
        'risk_warnings': [
            f'组合预估最大回撤{est_dd}%（目标≤{target_dd}%），{"回撤可控" if est_dd <= target_dd else "注意回撤略超目标"}',
            '投资有风险，过往业绩不预示未来表现，投资者需自行承担投资决策的风险',
            '市场剧烈波动时可能出现阶段性亏损，请根据自身风险承受能力合理配置',
            '建议定期（每季度）审视组合表现，根据市场变化和自身情况适时调整',
            '本推荐基于历史数据和量化模型，不构成投资建议，仅供参考',
        ],
    }


# ══════════════════════════════════════════════════════════════
# 回测引擎
# ══════════════════════════════════════════════════════════════

def run_backtest(portfolio_id, start_date, end_date):
    """运行历史回测"""
    nav_data, err = compute_portfolio_nav_data(portfolio_id, years=5)
    if err:
        return {'success': False, 'error': err}
    if not nav_data:
        return {'success': False, 'error': '无净值数据'}
    # 过滤时间段
    filtered = [d for d in nav_data if start_date <= d['nav_date'] <= end_date]
    if len(filtered) < 10:
        return {'success': False, 'error': '所选时间段数据不足（需要至少10个交易日）'}
    navs = [d['nav_value'] for d in filtered]
    nav_series = pd.Series(navs)
    total_return = (nav_series.iloc[-1] / nav_series.iloc[0] - 1) * 100
    days = len(nav_series)
    year_frac = days / 252
    annual_return = ((1 + total_return / 100) ** (1 / max(year_frac, 0.02)) - 1) * 100
    daily_returns = nav_series.pct_change().dropna()
    annual_volatility = daily_returns.std() * np.sqrt(252) * 100
    rf = 2.5
    sharpe = (annual_return - rf) / annual_volatility if annual_volatility > 0 else 0
    downside = daily_returns[daily_returns < 0]
    downside_vol = downside.std() * np.sqrt(252) * 100 if len(downside) > 0 else annual_volatility
    sortino = (annual_return - rf) / downside_vol if downside_vol > 0 else 0
    peak = nav_series.expanding().max()
    drawdown_series = (nav_series / peak - 1) * 100
    max_dd = float(drawdown_series.min())
    # 正收益概率
    prob_result = _calc_positive_prob(nav_series)
    # 收益分布
    monthly_returns = _calc_period_returns(nav_series, 'month')
    quarterly_returns = _calc_period_returns(nav_series, 'quarter')
    yearly_returns = _calc_period_returns(nav_series, 'year')
    # 回撤事件
    drawdown_events = _find_drawdown_events(drawdown_series)
    # 生成 Plotly 图表数据
    chart_data = _build_backtest_chart(nav_data, start_date, end_date, filtered)
    result = {
        'success': True,
        'start_date': start_date,
        'end_date': end_date,
        'total_return': round(total_return, 2),
        'annual_return': round(annual_return, 2),
        'annual_volatility': round(annual_volatility, 2),
        'sharpe_ratio': round(sharpe, 2),
        'sortino_ratio': round(sortino, 2),
        'max_drawdown': round(max_dd, 2),
        'positive_prob': prob_result,
        'monthly_returns': monthly_returns,
        'quarterly_returns': quarterly_returns,
        'yearly_returns': yearly_returns,
        'drawdown_events': drawdown_events,
        'chart_data': chart_data,
        'nav_data': filtered[-500:],
    }
    # 缓存
    try:
        _db_execute(
            "INSERT INTO portfolio_backtest_cache (portfolio_id, start_date, end_date, backtest_data) "
            "VALUES (%s, %s, %s, %s) ON DUPLICATE KEY UPDATE backtest_data=%s",
            (portfolio_id, start_date, end_date, json.dumps(result, ensure_ascii=False),
             json.dumps(result, ensure_ascii=False)), fetch=False)
    except Exception:
        try:
            _db_execute(
                "INSERT OR REPLACE INTO portfolio_backtest_cache (portfolio_id, start_date, end_date, backtest_data) "
                "VALUES (?, ?, ?, ?)",
                (portfolio_id, start_date, end_date, json.dumps(result, ensure_ascii=False)), fetch=False)
        except Exception:
            pass
    return result


def _calc_positive_prob(nav_series):
    """计算不同持有期的正收益概率"""
    periods = {'1个月': 21, '3个月': 63, '6个月': 126, '1年': 252, '2年': 504, '3年': 756}
    result = {}
    for label, window in periods.items():
        if len(nav_series) <= window:
            result[label] = None
            continue
        positive = 0
        total = 0
        for i in range(len(nav_series) - window):
            ret = nav_series.iloc[i + window] / nav_series.iloc[i] - 1
            if ret > 0:
                positive += 1
            total += 1
        result[label] = round(positive / total * 100, 1) if total > 0 else None
    return result


def _calc_period_returns(nav_series, period='month'):
    """计算周期收益分布"""
    df = pd.DataFrame({'nav': nav_series})
    df.index = pd.to_datetime(df.index) if not isinstance(df.index, pd.DatetimeIndex) else df.index
    if period == 'month':
        grouped = df.resample('ME')
    elif period == 'quarter':
        grouped = df.resample('QE')
    else:
        grouped = df.resample('YE')
    returns = []
    for name, group in grouped:
        if len(group) >= 2:
            ret = (group['nav'].iloc[-1] / group['nav'].iloc[0] - 1) * 100
            label = name.strftime('%Y-%m') if period != 'year' else name.strftime('%Y')
            returns.append({'period': label, 'return': round(ret, 2)})
    return returns[-24:] if period != 'year' else returns[-5:]


def _find_drawdown_events(drawdown_series, threshold=-5):
    """找出超过阈值的回撤事件"""
    events = []
    in_drawdown = False
    start_idx = 0
    max_dd = 0
    for i, dd in enumerate(drawdown_series):
        if dd < threshold and not in_drawdown:
            in_drawdown = True
            start_idx = i
            max_dd = dd
        elif in_drawdown:
            if dd < max_dd:
                max_dd = dd
            if dd >= threshold or i == len(drawdown_series) - 1:
                if start_idx < len(drawdown_series):
                    events.append({
                        'start': str(drawdown_series.index[start_idx])[:10],
                        'end': str(drawdown_series.index[min(i, len(drawdown_series)-1)])[:10],
                        'max_drawdown': round(float(max_dd), 2),
                    })
                in_drawdown = False
    return events[-10:]


def _build_backtest_chart(nav_data, start_date, end_date, filtered):
    """构建回测图表的 Plotly 数据"""
    dates = [d['nav_date'] for d in filtered]
    navs = [d['nav_value'] for d in filtered]
    return {'dates': dates, 'navs': navs}


def run_stress_test(portfolio_id, scenario='2008'):
    """压力测试（模拟）"""
    scenarios = {
        '2008': {'name': '2008年金融危机', 'market_drop': -65, 'duration_months': 12},
        '2015': {'name': '2015年股灾', 'market_drop': -45, 'duration_months': 6},
        '2020': {'name': '2020年疫情冲击', 'market_drop': -15, 'duration_months': 3},
    }
    sc = scenarios.get(scenario, scenarios['2008'])
    nav_data, err = compute_portfolio_nav_data(portfolio_id, years=3)
    max_dd = 0
    if nav_data:
        navs = [d['nav_value'] for d in nav_data]
        nav_series = pd.Series(navs)
        peak = nav_series.expanding().max()
        drawdown = (nav_series / peak - 1) * 100
        max_dd = float(drawdown.min())
    beta = min(abs(max_dd) / 30, 1.5) if max_dd < 0 else 0.8
    estimated_loss = sc['market_drop'] * beta
    return {
        'success': True,
        'scenario': sc['name'],
        'market_drop': sc['market_drop'],
        'portfolio_beta': round(beta, 2),
        'estimated_loss': round(estimated_loss, 1),
        'duration_months': sc['duration_months'],
        'analysis': f"在{sc['name']}情景下，市场下跌{abs(sc['market_drop'])}%，"
                     f"组合预估回撤约{abs(estimated_loss):.1f}%（基于组合Beta={beta:.2f}）。"
                     f"建议在极端行情下保持冷静，避免恐慌性赎回。",
    }


def run_scenario_analysis(portfolio_id, scenario_type='bull'):
    """情景分析（模拟）"""
    scenarios = {
        'bull': {'name': '牛市场景（沪深300上涨30%）', 'market_return': 30, 'beta_multiplier': 1.2},
        'bear': {'name': '熊市场景（沪深300下跌30%）', 'market_return': -30, 'beta_multiplier': 1.1},
        'ranging': {'name': '震荡场景（沪深300波动±10%）', 'market_return': 0, 'beta_multiplier': 0.5},
    }
    sc = scenarios.get(scenario_type, scenarios['bull'])
    nav_data, err = compute_portfolio_nav_data(portfolio_id, years=1)
    beta = 0.8
    if nav_data and len(nav_data) > 20:
        navs = pd.Series([d['nav_value'] for d in nav_data])
        daily_returns = navs.pct_change().dropna()
        beta = max(0.2, min(daily_returns.std() / 0.015, 2.0))
    estimated_return = sc['market_return'] * beta * sc['beta_multiplier']
    return {
        'success': True,
        'scenario': sc['name'],
        'market_return': sc['market_return'],
        'portfolio_beta': round(beta, 2),
        'estimated_return': round(estimated_return, 1),
        'analysis': f"在{sc['name']}中，"
                     f"组合预估收益约{estimated_return:+.1f}%。"
                     f"建议{'维持仓位' if estimated_return > -5 else '适当减仓'}。",
    }


# ══════════════════════════════════════════════════════════════
# 导出与分享
# ══════════════════════════════════════════════════════════════

def export_portfolio_excel(portfolio_id):
    """导出组合为 Excel 文件"""
    portfolio = get_portfolio(portfolio_id)
    if not portfolio:
        return None, '组合不存在'
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None, 'openpyxl 未安装'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '组合概览'
    # 标题
    ws.merge_cells('A1:F1')
    ws['A1'] = f"基金组合: {portfolio['name']}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    # 基本信息
    info_rows = [
        ('风险等级', portfolio['risk_level']),
        ('组合标签', '、'.join(portfolio['tags']) if portfolio['tags'] else '无'),
        ('创建时间', str(portfolio['created_at'])[:10] if portfolio['created_at'] else ''),
        ('描述', portfolio['description'] or '无'),
    ]
    for i, (k, v) in enumerate(info_rows, 3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=str(v))
    # 核心指标
    metrics = portfolio.get('metrics', {})
    if metrics:
        ws.cell(row=8, column=1, value='核心指标').font = Font(size=13, bold=True)
        metric_keys = [
            ('总收益', 'total_return', '%'),
            ('年化收益', 'annual_return', '%'),
            ('年化波动率', 'annual_volatility', '%'),
            ('夏普比率', 'sharpe_ratio', ''),
            ('索提诺比率', 'sortino_ratio', ''),
            ('最大回撤', 'max_drawdown', '%'),
        ]
        for i, (label, key, unit) in enumerate(metric_keys, 9):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            val = metrics.get(key)
            ws.cell(row=i, column=2, value=f"{val}{unit}" if val is not None else 'N/A')
    # 持仓明细
    holdings = portfolio.get('holdings', [])
    if holdings:
        start_row = 17
        ws.cell(row=start_row, column=1, value='持仓明细').font = Font(size=13, bold=True)
        headers = ['序号', '基金代码', '基金名称', '基金类型', '持仓比例']
        for j, h in enumerate(headers):
            cell = ws.cell(row=start_row + 1, column=j + 1, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
            cell.font = Font(bold=True, color='F8FAFC')
        for i, h in enumerate(holdings):
            ws.cell(row=start_row + 2 + i, column=1, value=i + 1)
            ws.cell(row=start_row + 2 + i, column=2, value=h['fund_code'])
            ws.cell(row=start_row + 2 + i, column=3, value=h.get('fund_name', ''))
            ws.cell(row=start_row + 2 + i, column=4, value=h.get('fund_type', ''))
            ws.cell(row=start_row + 2 + i, column=5, value=f"{h['weight']}%")
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{portfolio['name']}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return output, filename


def generate_share_text(portfolio_id):
    """生成分享文本"""
    portfolio = get_portfolio(portfolio_id)
    if not portfolio:
        return {'success': False, 'error': '组合不存在'}
    metrics = portfolio.get('metrics', {})
    lines = [
        f"📊 基金组合: {portfolio['name']}",
        f"🎯 风险等级: {portfolio['risk_level']}",
        f"📈 年化收益: {metrics.get('annual_return', 'N/A')}%",
        f"📉 最大回撤: {metrics.get('max_drawdown', 'N/A')}%",
        f"⭐ 夏普比率: {metrics.get('sharpe_ratio', 'N/A')}",
        "",
        "持仓基金:",
    ]
    for h in portfolio.get('holdings', []):
        lines.append(f"  • {h.get('fund_name', h['fund_code'])} ({h['weight']}%)")
    lines.append("")
    lines.append(f"创建时间: {str(portfolio.get('created_at', ''))[:10]}")
    return {'success': True, 'share_text': '\n'.join(lines)}


# ══════════════════════════════════════════════════════════════
# Flask 路由注册
# ══════════════════════════════════════════════════════════════

def _check_and_respond():
    """统一访问检查（用户级）：登录 + 订阅"""
    try:
        # 1. 检查登录
        if 'user_id' not in session:
            return jsonify({'success': False, 'error': '请先登录', 'code': 'login_required'})
        # 2. 检查系统收费开关
        settings = get_settings()
        if not settings.get('payment_enabled'):
            return None
        if settings.get('user_free_access'):
            return None
        # 3. 检查用户订阅
        from auth_manager import _get_user_active_subscription
        sub = _get_user_active_subscription(session['user_id'])
        if sub:
            return None
        # 4. 检查免费试用
        trial_start = settings.get('trial_start_date')
        if trial_start:
            try:
                ts = datetime.strptime(str(trial_start)[:10], '%Y-%m-%d')
                days_used = (datetime.now() - ts).days
                if days_used <= settings.get('free_trial_days', 7):
                    return None
            except Exception:
                pass
        return jsonify({'success': False, 'error': '需要订阅才能使用此功能', 'code': 'subscription_required'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e), 'code': 'error'})


def register_routes(app):
    """注册所有组合管理路由到 Flask app"""
    _init_deps()

    # ── 设置与支付 ──
    @app.route('/api/portfolio/settings', methods=['GET', 'PUT'])
    def portfolio_settings():
        if request.method == 'GET':
            return jsonify({'success': True, 'data': get_settings()})
        else:
            data = request.get_json() or {}
            return jsonify({'success': True, 'data': update_settings(data)})

    @app.route('/api/portfolio/access-status')
    def portfolio_access_status():
        allowed, reason = check_access()
        settings = get_settings()
        trial_days_left = 0
        if settings.get('trial_start_date'):
            try:
                ts = datetime.strptime(str(settings['trial_start_date'])[:10], '%Y-%m-%d')
                used = (datetime.now() - ts).days
                trial_days_left = max(0, settings.get('free_trial_days', 7) - used)
            except Exception:
                pass
        return jsonify({
            'success': True,
            'allowed': allowed,
            'reason': reason,
            'payment_enabled': settings['payment_enabled'],
            'trial_days_left': trial_days_left,
        })

    @app.route('/api/portfolio/subscribe', methods=['POST'])
    def portfolio_subscribe():
        data = request.get_json() or {}
        plan_type = data.get('plan_type', 'monthly')
        result = create_subscription(plan_type)
        return jsonify(result)

    @app.route('/api/portfolio/payment-history')
    def portfolio_payment_history():
        return jsonify({'success': True, 'data': get_payment_history()})

    # ── 风险评估 ──
    @app.route('/api/portfolio/risk/status')
    def portfolio_risk_status():
        uid = session.get('user_id')
        return jsonify({'success': True, 'data': get_risk_status(uid)})

    @app.route('/api/portfolio/risk/questions')
    def portfolio_risk_questions():
        return jsonify({'success': True, 'data': RISK_QUESTIONS})

    @app.route('/api/portfolio/risk/submit', methods=['POST'])
    def portfolio_risk_submit():
        uid = session.get('user_id')
        data = request.get_json() or {}
        answers = data.get('answers', [])
        result = submit_risk_assessment(answers, uid)
        return jsonify(result)

    @app.route('/api/portfolio/risk/override', methods=['PUT'])
    def portfolio_risk_override():
        uid = session.get('user_id')
        data = request.get_json() or {}
        new_dd = float(data.get('max_drawdown', 0))
        accepted = data.get('accepted', False)
        result = override_max_drawdown(new_dd, accepted, user_id=uid)
        return jsonify(result)

    @app.route('/api/portfolio/risk/history')
    def portfolio_risk_history():
        uid = session.get('user_id')
        return jsonify({'success': True, 'data': get_risk_history(uid)})

    # ── 组合 CRUD ──
    @app.route('/api/portfolio/list')
    def portfolio_list():
        access = _check_and_respond()
        if access: return access
        uid = session.get('user_id')
        return jsonify({'success': True, 'data': list_portfolios(uid)})

    @app.route('/api/portfolio/create', methods=['POST'])
    def portfolio_create():
        access = _check_and_respond()
        if access: return access
        uid = session.get('user_id')
        data = request.get_json() or {}
        result = create_portfolio(data, uid)
        return jsonify(result)

    @app.route('/api/portfolio/<int:portfolio_id>')
    def portfolio_detail(portfolio_id):
        access = _check_and_respond()
        if access: return access
        uid = session.get('user_id')
        result = get_portfolio(portfolio_id, uid)
        if result is None:
            return jsonify({'success': False, 'error': '组合不存在'})
        return jsonify({'success': True, 'data': result})

    @app.route('/api/portfolio/<int:portfolio_id>', methods=['PUT'])
    def portfolio_update(portfolio_id):
        access = _check_and_respond()
        if access: return access
        uid = session.get('user_id')
        data = request.get_json() or {}
        result = update_portfolio(portfolio_id, data, uid)
        return jsonify(result)

    @app.route('/api/portfolio/<int:portfolio_id>', methods=['DELETE'])
    def portfolio_delete(portfolio_id):
        access = _check_and_respond()
        if access: return access
        uid = session.get('user_id')
        result = delete_portfolio(portfolio_id, uid)
        return jsonify(result)

    @app.route('/api/portfolio/<int:portfolio_id>/duplicate', methods=['POST'])
    def portfolio_duplicate(portfolio_id):
        access = _check_and_respond()
        if access: return access
        uid = session.get('user_id')
        result = duplicate_portfolio(portfolio_id, uid)
        return jsonify(result)

    # ── 持仓管理 ──
    @app.route('/api/portfolio/<int:portfolio_id>/holdings', methods=['GET'])
    def portfolio_holdings(portfolio_id):
        access = _check_and_respond()
        if access: return access
        uid = session.get('user_id')
        # Verify ownership
        portfolio = get_portfolio(portfolio_id, uid)
        if portfolio is None:
            return jsonify({'success': False, 'error': '组合不存在或无权访问'})
        return jsonify({'success': True, 'data': get_holdings(portfolio_id)})

    @app.route('/api/portfolio/<int:portfolio_id>/holdings', methods=['POST'])
    def portfolio_add_holdings(portfolio_id):
        access = _check_and_respond()
        if access: return access
        uid = session.get('user_id')
        # Verify ownership
        portfolio = get_portfolio(portfolio_id, uid)
        if portfolio is None:
            return jsonify({'success': False, 'error': '组合不存在或无权访问'})
        data = request.get_json() or {}
        fund_codes = data.get('fund_codes', [])
        weights = data.get('weights', None)
        result = add_holdings(portfolio_id, fund_codes, weights)
        return jsonify(result)

    @app.route('/api/portfolio/<int:portfolio_id>/holdings/<int:holding_id>', methods=['PUT'])
    def portfolio_update_holding(portfolio_id, holding_id):
        access = _check_and_respond()
        if access: return access
        uid = session.get('user_id')
        # Verify ownership
        portfolio = get_portfolio(portfolio_id, uid)
        if portfolio is None:
            return jsonify({'success': False, 'error': '组合不存在或无权访问'})
        data = request.get_json() or {}
        weight = float(data.get('weight', 0))
        result = update_holding(holding_id, weight)
        return jsonify(result)

    @app.route('/api/portfolio/<int:portfolio_id>/holdings/<int:holding_id>', methods=['DELETE'])
    def portfolio_remove_holding(portfolio_id, holding_id):
        access = _check_and_respond()
        if access: return access
        uid = session.get('user_id')
        # Verify ownership
        portfolio = get_portfolio(portfolio_id, uid)
        if portfolio is None:
            return jsonify({'success': False, 'error': '组合不存在或无权访问'})
        result = remove_holding(holding_id)
        return jsonify(result)

    # ── 搜索与建议 ──
    @app.route('/api/portfolio/fund-search')
    def portfolio_fund_search():
        keyword = request.args.get('keyword', '')
        results = search_funds_for_portfolio(keyword)
        return jsonify({'success': True, 'data': results})

    @app.route('/api/portfolio/suggested-funds')
    def portfolio_suggested_funds():
        risk_type = request.args.get('risk_type', None)
        return jsonify({'success': True, 'data': get_suggested_funds(risk_type)})

    @app.route('/api/portfolio/name-generator')
    def portfolio_name_generator():
        risk_level = request.args.get('risk_level', '平衡型')
        style = request.args.get('style', None)
        return jsonify({'success': True, 'data': generate_portfolio_names(risk_level, style)})

    # ── 推荐 ──
    @app.route('/api/portfolio/recommend')
    def portfolio_recommend():
        risk_level = request.args.get('risk_level', '平衡型')
        return jsonify({'success': True, 'data': generate_recommendation(risk_level)})

    @app.route('/api/portfolio/recommend/create', methods=['POST'])
    def portfolio_recommend_create():
        access = _check_and_respond()
        if access: return access
        uid = session.get('user_id')
        data = request.get_json() or {}
        rec = data.get('recommendation', {})
        funds = [(f['fund_code'], f['weight']) for f in rec.get('funds', [])]
        result = create_portfolio({
            'mode': 'recommend',
            'name': rec.get('name', f"{rec.get('risk_level', '平衡')}推荐组合"),
            'risk_level': rec.get('risk_level', '平衡型'),
            'fund_codes': [f[0] for f in funds],
            'fund_weights': [f[1] for f in funds],
        }, uid)
        return jsonify(result)

    # ── 净值与图表 ──
    @app.route('/api/portfolio/<int:portfolio_id>/nav')
    def portfolio_nav(portfolio_id):
        uid = session.get('user_id')
        # Verify ownership
        portfolio = get_portfolio(portfolio_id, uid)
        if portfolio is None:
            return jsonify({'success': False, 'error': '组合不存在或无权访问'})
        years = int(request.args.get('years', 3))
        nav_data, err = compute_portfolio_nav_data(portfolio_id, years=years)
        if err:
            return jsonify({'success': False, 'error': err})
        return jsonify({'success': True, 'data': nav_data})

    @app.route('/api/portfolio/<int:portfolio_id>/metrics')
    def portfolio_metrics(portfolio_id):
        uid = session.get('user_id')
        # Verify ownership
        portfolio = get_portfolio(portfolio_id, uid)
        if portfolio is None:
            return jsonify({'success': False, 'error': '组合不存在或无权访问'})
        metrics = _compute_portfolio_metrics(portfolio_id)
        return jsonify({'success': True, 'data': metrics})

    @app.route('/api/portfolio/<int:portfolio_id>/analysis')
    def portfolio_analysis(portfolio_id):
        uid = session.get('user_id')
        # Verify ownership
        portfolio = get_portfolio(portfolio_id, uid)
        if portfolio is None:
            return jsonify({'success': False, 'error': '组合不存在或无权访问'})
        analysis = compute_portfolio_analysis(portfolio_id)
        return jsonify({'success': True, 'data': analysis})

    # ── 回测 ──
    @app.route('/api/portfolio/<int:portfolio_id>/backtest')
    def portfolio_backtest(portfolio_id):
        uid = session.get('user_id')
        # Verify ownership
        portfolio = get_portfolio(portfolio_id, uid)
        if portfolio is None:
            return jsonify({'success': False, 'error': '组合不存在或无权访问'})
        start_date = request.args.get('start_date', (date.today() - timedelta(days=365)).strftime('%Y-%m-%d'))
        end_date = request.args.get('end_date', date.today().strftime('%Y-%m-%d'))
        result = run_backtest(portfolio_id, start_date, end_date)
        return jsonify(result)

    @app.route('/api/portfolio/<int:portfolio_id>/backtest/stress-test')
    def portfolio_stress_test(portfolio_id):
        uid = session.get('user_id')
        # Verify ownership
        portfolio = get_portfolio(portfolio_id, uid)
        if portfolio is None:
            return jsonify({'success': False, 'error': '组合不存在或无权访问'})
        scenario = request.args.get('scenario', '2008')
        result = run_stress_test(portfolio_id, scenario)
        return jsonify(result)

    @app.route('/api/portfolio/<int:portfolio_id>/backtest/scenario')
    def portfolio_scenario_analysis(portfolio_id):
        uid = session.get('user_id')
        # Verify ownership
        portfolio = get_portfolio(portfolio_id, uid)
        if portfolio is None:
            return jsonify({'success': False, 'error': '组合不存在或无权访问'})
        scenario_type = request.args.get('type', 'bull')
        result = run_scenario_analysis(portfolio_id, scenario_type)
        return jsonify(result)

    # ── 导出与分享 ──
    @app.route('/api/portfolio/<int:portfolio_id>/export')
    def portfolio_export(portfolio_id):
        uid = session.get('user_id')
        # Verify ownership
        portfolio = get_portfolio(portfolio_id, uid)
        if portfolio is None:
            return jsonify({'success': False, 'error': '组合不存在或无权访问'})
        from flask import send_file
        excel_data, filename = export_portfolio_excel(portfolio_id)
        if excel_data is None:
            return jsonify({'success': False, 'error': filename})
        return send_file(
            excel_data,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )

    @app.route('/api/portfolio/<int:portfolio_id>/share', methods=['POST'])
    def portfolio_share(portfolio_id):
        uid = session.get('user_id')
        # Verify ownership
        portfolio = get_portfolio(portfolio_id, uid)
        if portfolio is None:
            return jsonify({'success': False, 'error': '组合不存在或无权访问'})
        result = generate_share_text(portfolio_id)
        return jsonify(result)

    print("[portfolio] All routes registered")
